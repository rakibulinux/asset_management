# Copyright (c) 2026, Shayan and contributors
# For license information, please see license.txt

import io
import os
import re
from datetime import datetime

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cstr, flt, getdate, now_datetime
from frappe.utils.file_manager import get_file_path


REQUIRED_COLUMNS = [
    "asset_name",
    "item_name",
    "category",
    "location",
    "rfid_tag",
    "images",
]

COLUMN_ALIASES = {
    "asset name": "asset_name",
    "asset_name": "asset_name",
    "item name": "item_name",
    "item_name": "item_name",
    "category": "category",
    "asset category": "category",
    "asset_category": "category",
    "location": "location",
    "rfid tag": "rfid_tag",
    "rfid_tag": "rfid_tag",
    "rfid": "rfid_tag",
    "images": "images",
    "image": "images",
}


class AssetBulkImport(Document):
    @frappe.whitelist()
    def start_import(self):
        if self.status == "Importing":
            frappe.throw(_("This import is already running."))

        if not self.import_file:
            frappe.throw(_("Please upload an Excel or CSV file first."))

        for f in ("default_company", "default_purchase_date",
                  "default_available_for_use_date", "default_gross_purchase_amount"):
            if not self.get(f):
                frappe.throw(_("Default field is required: {0}").format(self.meta.get_label(f)))

        self.status = "Importing"
        self.rows_total = 0
        self.rows_success = 0
        self.rows_failed = 0
        self.import_log = ""
        self.save(ignore_permissions=True)
        frappe.db.commit()

        log_lines = []
        success = 0
        failed = 0

        try:
            rows = _read_rows(self.import_file)
        except Exception as e:
            self.status = "Failed"
            self.import_log = _("Could not read file: {0}").format(str(e))
            self.save(ignore_permissions=True)
            frappe.db.commit()
            return {"success": False, "message": str(e)}

        self.rows_total = len(rows)

        for index, row in enumerate(rows, start=2):
            try:
                _validate_row(row, index)
                asset_name = self._process_row(row)
                success += 1
                log_lines.append(_("Row {0}: created Asset '{1}'").format(index, asset_name))
            except Exception as e:
                failed += 1
                log_lines.append(_("Row {0}: ERROR - {1}").format(index, str(e)))
                frappe.log_error(
                    title="Asset Bulk Import row error",
                    message=f"Row {index}: {frappe.get_traceback()}",
                )

        self.rows_success = success
        self.rows_failed = failed
        if failed == 0 and success > 0:
            self.status = "Success"
        elif success == 0:
            self.status = "Failed"
        else:
            self.status = "Partial Success"

        self.import_log = "\n".join(log_lines)
        self.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "success": True,
            "rows_total": self.rows_total,
            "rows_success": success,
            "rows_failed": failed,
        }

    def _process_row(self, row):
        asset_name = cstr(row["asset_name"]).strip()
        item_name = cstr(row["item_name"]).strip()
        category = cstr(row["category"]).strip()
        location = cstr(row["location"]).strip()
        rfid_tag = cstr(row["rfid_tag"]).strip()
        images_raw = cstr(row["images"]).strip()

        if frappe.db.exists("Asset", {"asset_name": asset_name}):
            frappe.throw(_("Asset with name '{0}' already exists").format(asset_name))

        if frappe.db.exists("Asset", {"rfid_tag": rfid_tag}):
            frappe.throw(_("Asset with RFID tag '{0}' already exists").format(rfid_tag))

        item_group = _ensure_item_group(category)
        asset_category = _ensure_asset_category(category, item_group)
        _ensure_location(location)
        item_code = _ensure_item(item_name, item_group, asset_category)

        asset = frappe.new_doc("Asset")
        asset.update({
            "asset_name": asset_name,
            "item_code": item_code,
            "asset_category": asset_category,
            "location": location,
            "rfid_tag": rfid_tag,
            "company": self.default_company,
            "purchase_date": getdate(self.default_purchase_date),
            "available_for_use_date": getdate(self.default_available_for_use_date),
            "gross_purchase_amount": flt(self.default_gross_purchase_amount),
            "is_existing_asset": 1,
            "calculate_depreciation": 0,
            "asset_quantity": 1,
        })

        for url in _split_image_urls(images_raw):
            asset.append("images", {"image": url})

        asset.flags.ignore_permissions = True
        asset.insert()

        return asset.name


def _read_rows(file_url):
    file_path = _resolve_file_path(file_url)
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(file_path)
    if ext == ".csv":
        return _read_csv(file_path)
    raise frappe.ValidationError(
        _("Unsupported file type '{0}'. Use .xlsx or .csv").format(ext)
    )


def _resolve_file_path(file_url):
    if not file_url:
        raise frappe.ValidationError(_("No file attached"))
    return get_file_path(file_url)


def _read_xlsx(path):
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    ws = wb.active

    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [_normalize_header(h) for h in row]
            _check_headers(headers)
            continue
        if all(cell is None or cstr(cell).strip() == "" for cell in row):
            continue
        rows.append(_row_to_dict(headers, row))

    wb.close()
    return rows


def _read_csv(path):
    import csv

    with io.open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = []
        rows = []
        for i, row in enumerate(reader):
            if i == 0:
                headers = [_normalize_header(h) for h in row]
                _check_headers(headers)
                continue
            if all(cstr(cell).strip() == "" for cell in row):
                continue
            rows.append(_row_to_dict(headers, row))
        return rows


def _normalize_header(h):
    if h is None:
        return ""
    key = cstr(h).strip().lower()
    return COLUMN_ALIASES.get(key, key)


def _check_headers(headers):
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        labels = {
            "asset_name": "Asset Name",
            "item_name": "Item Name",
            "category": "Category",
            "location": "Location",
            "rfid_tag": "RFID Tag",
            "images": "Images",
        }
        pretty = [labels[m] for m in missing]
        raise frappe.ValidationError(
            _("Missing required columns: {0}").format(", ".join(pretty))
        )


def _row_to_dict(headers, row):
    out = {}
    for i, key in enumerate(headers):
        if not key:
            continue
        out[key] = row[i] if i < len(row) else None
    return out


def _validate_row(row, index):
    for col in REQUIRED_COLUMNS:
        value = cstr(row.get(col, "")).strip()
        if not value:
            raise frappe.ValidationError(
                _("Column '{0}' is empty").format(col)
            )


def _split_image_urls(raw):
    if not raw:
        return []
    parts = re.split(r"[,|\n;]+", cstr(raw))
    return [p.strip() for p in parts if p.strip()]


def _ensure_item_group(category_name):
    name = cstr(category_name).strip()
    if not name:
        frappe.throw(_("Category cannot be empty"))

    if frappe.db.exists("Item Group", name):
        return name

    parent = "All Item Groups" if frappe.db.exists("Item Group", "All Item Groups") else None
    doc = frappe.new_doc("Item Group")
    doc.item_group_name = name
    if parent:
        doc.parent_item_group = parent
        doc.is_group = 0
    doc.flags.ignore_permissions = True
    doc.insert()
    return doc.name

def _get_account_by_type(company, account_type):
    return frappe.db.get_value(
        "Account",
        {
            "company": company,
            "account_type": account_type,
            "is_group": 0,
            "disabled": 0,
        },
        "name",
    )

def _ensure_asset_category(category_name, item_group):
    name = cstr(category_name).strip()
    if not name:
        frappe.throw(_("Category cannot be empty"))

    company = (
        frappe.defaults.get_user_default("Company")
        or frappe.db.get_single_value("Global Defaults", "default_company")
    )

    if not company:
        frappe.throw(_("Default Company is not set"))

    fixed_asset_account = _get_account_by_type(company, "Fixed Asset")

    if not fixed_asset_account:
        frappe.throw(
            _("No Fixed Asset account found for company '{0}'. Please create an Account with Account Type 'Fixed Asset'.")
            .format(company)
        )

    if frappe.db.exists("Asset Category", name):
        doc = frappe.get_doc("Asset Category", name)

        existing_row = None
        for row in doc.accounts:
            if row.company_name == company:
                existing_row = row
                break

        if existing_row:
            if not existing_row.fixed_asset_account:
                existing_row.fixed_asset_account = fixed_asset_account
        else:
            doc.append("accounts", {
                "company_name": company,
                "fixed_asset_account": fixed_asset_account,
            })

        doc.flags.ignore_permissions = True
        doc.save()
        return doc.name

    doc = frappe.new_doc("Asset Category")
    doc.asset_category_name = name
    doc.enable_cwip_accounting = 0

    doc.append("accounts", {
        "company_name": company,
        "fixed_asset_account": fixed_asset_account,
    })

    doc.flags.ignore_permissions = True
    doc.insert()
    return doc.name

def _ensure_location(location_name):
    name = cstr(location_name).strip()
    if not name:
        frappe.throw(_("Location cannot be empty"))

    if frappe.db.exists("Location", name):
        return name

    doc = frappe.new_doc("Location")
    doc.location_name = name
    doc.is_group = 0
    doc.flags.ignore_permissions = True
    doc.insert()
    return doc.name


def _ensure_item(item_name, item_group, asset_category):
    name = cstr(item_name).strip()
    if not name:
        frappe.throw(_("Item Name cannot be empty"))

    existing = frappe.db.get_value(
        "Item",
        {"item_name": name, "is_fixed_asset": 1},
        "name",
    )
    if existing:
        return existing

    code = _make_item_code(name)

    doc = frappe.new_doc("Item")
    doc.item_code = code
    doc.item_name = name
    doc.item_group = item_group
    doc.is_fixed_asset = 1
    doc.is_stock_item = 0
    doc.asset_category = asset_category
    doc.auto_create_assets = 0
    doc.include_item_in_manufacturing = 0
    doc.flags.ignore_permissions = True
    doc.insert()
    return doc.name


def _make_item_code(item_name):
    base = re.sub(r"\s+", "-", cstr(item_name).strip())
    base = re.sub(r"[^A-Za-z0-9\-_]", "", base)[:120] or "ITEM"

    if not frappe.db.exists("Item", base):
        return base

    suffix = 1
    while True:
        candidate = f"{base}-{suffix}"
        if not frappe.db.exists("Item", candidate):
            return candidate
        suffix += 1
